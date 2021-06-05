import csv, statsapi, datetime, PitcherClass, HitterClass, requests, json, TeamAverages, LeagueAverages, RotogrindersLineups
from pybaseball import playerid_reverse_lookup
from dateutil import tz
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from Colors import Colors

PLAYER_STATS_URL = "https://cdn.fangraphs.com/api/players/splits?playerid={}&position={}&season={}&split=&z=1614959387TEAM_CHANGE"
teamAvg = TeamAverages.TeamAverages()
leagueAvg = LeagueAverages.LeagueAverages()
rgl = RotogrindersLineups.RotogrindersLineups()
file = open("MissingPlayerIds.csv", "a+")
manualFill = True
parkFactors = json.loads(open("ParkFactors.json", "r").read())
socket = None
askQuestion = None

def setSocket(so):
    if so != None:
        socket = so

def setAskQuestionCallback(cb):
    global askQuestion
    askQuestion = cb

def cleanUp():
    file.close()

def filterTime(game, slateStart, compare):
    from_zone = tz.tzutc()
    to_zone = tz.tzlocal()

    dt = datetime.datetime.strptime(game['game_datetime'], "%Y-%m-%dT%H:%M:%SZ")
    dt = dt.replace(tzinfo=from_zone)
    local = dt.astimezone(to_zone)
    
    if compare == "before":
        if local.time() <= slateStart.time():
            return True
        else:
            return False
    elif compare == "after":
        if local.time() >= slateStart.time():
            return True
        else:
            return False

#Get games for either today or a specified day
def getDaysGames(mf, slateStart, compare, getImpliedTotals, date = None):
    global manualFill
    manualFill = mf

    if date == None:
        dt = datetime.datetime.now()
        
        day = ""
        if dt.day < 10:
            day = "0{}".format(dt.day)
        else:
            day = dt.day

        month = ""
        if dt.month < 10:
            month = "0{}".format(dt.month)
        else:
            month = dt.month
        
        year = dt.year
        
        date = "{}/{}/{}".format(month, day, year)
    
    games = statsapi.schedule(date)
    filteredGames = []
    for game in games:
        if filterTime(game, slateStart, compare):
            filteredGames.append(game)

    for game in filteredGames:
        if getImpliedTotals:
            game['impliedTotals'] = getTeamsImpliedTotals(game)

    return filteredGames

def getTeamsImpliedTotals(game):
    set = {}
    homeTotal = input("Please provide the implied total for {}:\n ".format(game['home_name']))
    awayTotal = input("Please provide the implied total for {}:\n ".format(game['away_name']))
    set['homeTotal'] = homeTotal
    set['awayTotal'] = awayTotal
    set['combinedTotal'] = homeTotal + awayTotal

    return set

def getGamebyID(gid):
    game = statsapi.schedule(game_id=gid)
    return game

def getTeamRoster(teamId):
    data = statsapi.get("team_roster", {"teamId": teamId})
    return data

def getPlayerInfo(name):
    data = statsapi.lookup_player(name)
    return data

def createHitters(pitcher, hitters):
    roster = pitcher.oppTeamRoster

    for player in roster:
        #Create hitter
        hitter = HitterClass.HitterClass(player, leagueAvg, pitcher)
        if hitter.position == "P":
            continue
        hitter.fid = getFangraphsId(hitter)
        if hitter.fid == None:
            continue
        hitter.stats = getMostRecentStats(hitter.fid, hitter.position)
        hitter.setOtherInformation(setHitterOtherInfo(pitcher.gameInfo, pitcher.home), rgl)
        hitter.handedness = getPitcherHandedness(statsapi.get('person', {'personId': str(hitter.pid)}))

        hitters.append(hitter)

def getGamesProbablePitchers(game, pitchers):
    global askQuestion
    homeRoster = getTeamRoster(game["home_id"])["roster"]
    awayRoster = getTeamRoster(game["away_id"])["roster"]
    pitcherHome = game["home_probable_pitcher"]
    pitcherAway = game["away_probable_pitcher"]
    homePitcherFound = False
    awayPitcherFound = False

    if pitcherHome == None or pitcherHome == "":
        pitcherHome = input("Please input the name of today's pitcher for the {}".format(game["home_name"]))
        #pitcherHome = askQuestion("Please input the name of today's pitcher for the {}".format(game["home_name"]))
    if pitcherAway == None or pitcherAway == "":
        pitcherAway = input("Please input the name of today's pitcher for the {}".format(game["away_name"]))
        #pitcherAway = askQuestion("Please input the name of today's pitcher for the {}".format(game["away_name"]))

    try:
        parkFactorsForVenue = parkFactors[game['venue_name']]
    except Exception as e:
        print("\n\n\nError in getting Park Factors: {}\nCould not get Park Factors for {}\n\n\n".format(str(e), game['venue_name']))
        parkFactorsForVenue = {"runs": 0, "hr": 0}

    print("{} vs {}".format(pitcherHome, pitcherAway))

    pitcher = None
    if pitcherHome != '':
        for p in homeRoster:
            if pitcherHome in p["person"]["fullName"]:
                print("Found {} with pid: {}".format(pitcherHome, p["person"]["id"]))
                homePitcherFound = True
                pitcher = PitcherClass.PitcherClass(p, teamAvg, leagueAvg, awayRoster, game, parkFactorsForVenue)
                pitcher.fid = getFangraphsId(pitcher)
                if pitcher.fid == None:
                    continue
                pitcher.stats = getMostRecentStats(pitcher.fid, 'P')
                pitcher.setOtherInformation(setOtherInfo(game, "home"), rgl)
                pitcher.handedness = getPitcherHandedness(statsapi.get('person', {'personId': str(pitcher.pid)}))
                if pitcher.stats['vsL'] != None and pitcher.stats['vsR'] != None:
                    #pitcher.assessSelf()
                    pitchers.append(pitcher)
                else:
                    pitcher.fid = askUserForFID(pitcher.name)
                    if pitcher.fid == None:
                        continue
                    pitcher.stats = getMostRecentStats(pitcher.fid, 'P')
                    if pitcher.stats['vsL'] != None and pitcher.stats['vsR'] != None:
                        #pitcher.assessSelf()
                        pitchers.append(pitcher)
        if not homePitcherFound:
            try:
                player = statsapi.lookup_player(pitcherHome)[0]
                if player != None or player['id'] != None:
                    data = {
                        'person': {
                            'id': player['id'],
                            'fullName': pitcherHome
                        },
                        'position': {
                            'abbreviation': player['primaryPosition']['abbreviation']
                        }
                    }
                    pitcher = PitcherClass.PitcherClass(data, teamAvg, leagueAvg, awayRoster, game, parkFactorsForVenue)
                    pitcher.fid = getFangraphsId(pitcher, manualFill)
                    if pitcher.fid != None:
                        pitcher.stats = getMostRecentStats(pitcher.fid, 'P')
                        pitcher.setOtherInformation(setOtherInfo(game, "home"), rgl)
                        pitcher.handedness = getPitcherHandedness(statsapi.get('person', {'personId': str(pitcher.pid)}))
                        if pitcher.stats['vsL'] != None and pitcher.stats['vsR'] != None:
                            #pitcher.assessSelf()
                            pitchers.append(pitcher)
                        else:
                            pitcher.fid = askUserForFID(pitcher.name)
                            if pitcher.fid == None:
                                pitcher.stats = getMostRecentStats(pitcher.fid, 'P')
                                if pitcher.stats['vsL'] != None and pitcher.stats['vsR'] != None:
                                    #pitcher.assessSelf()
                                    pitchers.append(pitcher)
            except:
                print("\n\Couldn't find home pitcher {}".format(pitcherHome))
            
    if pitcherAway != '':
        for p in awayRoster:
            if pitcherAway in p["person"]["fullName"]:
                print("Found {} with pid: {}".format(pitcherAway, p["person"]["id"]))
                awayPitcherFound = True
                pitcher = PitcherClass.PitcherClass(p, teamAvg, leagueAvg, homeRoster, game, parkFactorsForVenue)
                pitcher.fid = getFangraphsId(pitcher)
                if pitcher.fid == None:
                    continue
                pitcher.stats = getMostRecentStats(pitcher.fid, 'P')
                pitcher.setOtherInformation(setOtherInfo(game, "away"), rgl)
                pitcher.handedness = getPitcherHandedness(statsapi.get('person', {'personId': str(pitcher.pid)}))
                if pitcher.stats['vsL'] != None and pitcher.stats['vsR'] != None:
                    #pitcher.assessSelf()
                    pitchers.append(pitcher)
                else:
                    pitcher.fid = askUserForFID(pitcher.name)
                    if pitcher.fid == None:
                        continue
                    pitcher.stats = getMostRecentStats(pitcher.fid, 'P')
                    if pitcher.stats['vsL'] != None and pitcher.stats['vsR'] != None:
                        #pitcher.assessSelf()
                        pitchers.append(pitcher)
        if not awayPitcherFound:
            try:
                player = statsapi.lookup_player(pitcherAway)[0]
                if player != None or player['id'] != None:
                    data = {
                        'person': {
                            'id': player['id'],
                            'fullName': pitcherAway
                        },
                        'position': {
                            'abbreviation': player['primaryPosition']['abbreviation']
                        }
                    }
                    pitcher = PitcherClass.PitcherClass(data, teamAvg, leagueAvg, awayRoster, game, parkFactorsForVenue)
                    pitcher.fid = getFangraphsId(pitcher, manualFill)
                    if pitcher.fid != None:
                        pitcher.stats = getMostRecentStats(pitcher.fid, 'P')
                        pitcher.setOtherInformation(setOtherInfo(game, "home"), rgl)
                        pitcher.handedness = getPitcherHandedness(statsapi.get('person', {'personId': str(pitcher.pid)}))
                        if pitcher.stats['vsL'] != None and pitcher.stats['vsR'] != None:
                            #pitcher.assessSelf()
                            pitchers.append(pitcher)
                        else:
                            pitcher.fid = askUserForFID(pitcher.name)
                            if pitcher.fid == None:
                                pitcher.stats = getMostRecentStats(pitcher.fid, 'P')
                                if pitcher.stats['vsL'] != None and pitcher.stats['vsR'] != None:
                                    #pitcher.assessSelf()
                                    pitchers.append(pitcher)
            except:
                print("\nCould not find away Pitcher {}\n".format(pitcherAway))

def assessPitchers(pitchers):
    totalInnings = 0
    for pitcher in pitchers:
        totalInnings += pitcher.stats['vsL']['IP'] + pitcher.stats['vsR']['IP']
    
    avgIP = totalInnings / len(pitchers)

    for pitcher in pitchers:
        pitcher.assessSelf(avgIP)

def assessHitters(hitters):
    avgPA = 0
    count = 0
    for hitter in hitters:
        if hitter.position != 'P':
            vsLPA = 0
            try:
                vsLPA = hitter.stats['vsL']['PA']
            except:
                vsLPA = 0
            vsRPA = 0
            try:
                vsLPA = hitter.stats['vsR']['PA']
            except:
                vsLPA = 0
            avgPA +=  vsLPA + vsRPA
            count += 1
    avgPA = avgPA / count

    for hitter in hitters:
        if hitter.position != "P":
            #hitter.assessSelf(avgPA)
            hitter.assessSelfV2(avgPA)

def setHitterOtherInfo(data, location):
    info = {
        'oppTeamId': data['away_id'] if location == "away" else data['home_id'],
        'oppTeamName': data['away_name'] if location == "away" else data['home_name'],
        'teamId': data['home_id'] if location == "away" else data['away_id'],
        'teamName': data['home_name'] if location == "away" else data['away_name'],
        'stadiumId': data['venue_id'],
        'stadiumName': data['venue_name'],
    }

    return info

def setOtherInfo(data, location):
    info = {
        'oppTeamId': data['away_id'] if location == "home" else data['home_id'],
        'oppTeamName': data['away_name'] if location == "home" else data['home_name'],
        'teamId': data['home_id'] if location == "home" else data['away_id'],
        'teamName': data['home_name'] if location == "home" else data['away_name'],
        'stadiumId': data['venue_id'],
        'stadiumName': data['venue_name'],
        'homeOrAway': location
    }

    return info

def getPitcherHandedness(data):
    player = data['people'][0]
    handedness = player['pitchHand']['code']
    
    return handedness

def getFangraphsId(player, retrying = None):
    data = playerid_reverse_lookup([player.pid], "mlbam")
    
    fid = None
    try:
        fid = data.at[0, "key_fangraphs"]
    except:
        print("Couldn't get fangraphs ID for {}".format(player.name))
        if retrying == None:
            return checkInCsvRecords(player.pid, player.name)

    if fid == -1 and retrying == None:
        return checkInCsvRecords(player.pid, player.name)

    return fid

def checkInCsvRecords(pid, name):
    file.seek(0)
    csv_reader = csv.reader(file, delimiter=',')
    line_count = 0
    fid = None
    for row in csv_reader:
        if line_count > 0:
            try:
                if row[0] == str(pid):
                    fid = row[1]
                    return fid
            except:
                pass
        line_count += 1
    
    #Player Not Found
    fid = askUserForFID(name)
    if fid == None or fid == "":
        return None

    file.write("{},{},\n".format(pid, fid))
    return fid

def askUserForFID(name):
    if not manualFill:
        return None

    fid = input("Please provide {}'s confirmed Fangraphs ID:\n".format(name))
    #fid = askQuestion("Please provide {}'s confirmed Fangraphs ID:\n".format(name))

    if fid == None or fid == "":
        print("Returning None")
        return None
    else:
        return fid

def getMostRecentStats(pid, pos):
    d = datetime.datetime.now()
    year = d.year
    statsL = None
    statsR = None
    offset = 0
    while (offset <= 2 and statsL == None and statsR == None):
        adjustedYear = year - offset

        #if adjustedYear == 2021:
        #    offset += 1
        #    continue
        #TODO: REMOVE WHEN 2021 SAMPLE SIZE IS LARGE ENOUGH

        url = PLAYER_STATS_URL.format(pid, pos, adjustedYear)
        data = requests.get(url)
        data = json.loads(data.text)
        #print("Data from {}".format(url))
        #print(data)

        if (len(data) == 0):
            offset += 1
            continue

        for obj in data:
            try:
                if obj['Split'] == "vs L":
                    statsL = obj
                    statsL['season'] = obj['Season']
                elif obj['Split'] == "vs R":
                    statsR = obj
                    statsR['season'] = obj['Season']
            except:
                pass

    #Get Career Stats
    careerStatsL = None
    careerStatsR = None
    url = PLAYER_STATS_URL.format(pid, pos, 0)
    data = requests.get(url)
    data = json.loads(data.text)

    for obj in data:
        try:
            if obj['Split'] == "vs L":
                careerStatsL = obj
                careerStatsL['season'] = obj['Season']
            elif obj['Split'] == "vs R":
                careerStatsR = obj
                careerStatsR['season'] = obj['Season']
        except:
            pass

    stats = { 'vsL': statsL, 'vsR': statsR, 'careerVsL': careerStatsL, 'careerVsR': careerStatsR }
    return stats

def sortPitchers(pitchers):
    pitchers.sort(key=lambda x: x.overall, reverse=True)
    pSize = len(pitchers)
    pitcherSet = {"use": [], "target": []}
    if pSize == 0:
        print("No probable pitchers")
        return pitcherSet
    elif pSize == 1:
        print("Only 1 available pitcher")
        pitcherSet = { "use": [pitchers[0]], "target": [pitchers[0]]}
        return pitcherSet
    
    for x in range(0, pSize):
        pitcher = pitchers[x]
        if (x / (pSize - 1)) >= .75:
            pitcherSet["use"].append(pitcher)
        else:
            pitcherSet["target"].append(pitcher)
    
    return pitcherSet

def getPlayerSalaries(players):
    FDFile = open("FDPlayerList.csv", "r")
    FDFile.seek(0)
    result = []

    for player in players:
        name = player.name
        teamKey = teamAvg.getFanduelTeamKey(player.teamName)

        lineCount = 0
        csv_reader = csv.reader(FDFile, delimiter=',')
        for row in csv_reader:
            if lineCount > 0:
                try:
                    csvName = "{} {}".format(row[2], row[4])
                    if name == csvName and teamKey == row[9]:
                        player.salary = row[7]
                        player.FDPos = row[1]
                        result.append(player)
                except:
                    player.salary = -1000
            lineCount += 1
        
        FDFile.seek(0)
    
    FDFile.close()
    return result

def getStacks(players):
    stacks = {}
    addedTeams = []

    for player in players:
        if player.overall <= 0.0:
            continue

        team = player.teamName
        if team not in addedTeams:
            addedTeams.append(team)
            stacks[team] = {}
            stacks[team]['totalOverall'] = 0.0
            stacks[team]['playerList'] = []

        stacks[team]['totalOverall'] += player.overall
        stacks[team]['playerList'].append(player)

    result = []
    for team in addedTeams:
        var = {
            "team": team,
            "averageOverall": stacks[team]['totalOverall'] / len(stacks[team]['playerList']),
            "players": stacks[team]['playerList']
        }
        result.append(var)
    result.sort(key=lambda x: x['averageOverall'], reverse=True)
    return result

def writeSummary(players, pitchers, hrList):
    summaryFile = open("Summary.txt", "w")

    pSet = sortPitchers(pitchers)

    output = "Pitchers to use:\n"
    for pitcher in pSet["use"]:
        output += "\t{} ${} [{}] vs {} - Overall: {} K% Average: {} K% vs L: {} K% vs R: {}\n".format(pitcher.name, pitcher.salary, pitcher.teamName, pitcher.oppTeamName, round(pitcher.overall, 2), round(pitcher.kRate['avg'], 2), round(pitcher.kRate['vsL'], 2), round(pitcher.kRate['vsR'], 2))
    output += "\nPitchers to target:\n"
    for pitcher in pSet["target"]:
        output += "\t{} ${} [{}] vs {} - Overall: {} K% Average: {} K% vs L: {} K% vs R: {}\n".format(pitcher.name, pitcher.salary, pitcher.teamName, pitcher.oppTeamName, round(pitcher.overall, 2), round(pitcher.kRate['avg'], 2), round(pitcher.kRate['vsL'], 2), round(pitcher.kRate['vsR'], 2))
    output += "\nCatchers to use:\n"
    for p in players['C']: 
        output += "\t{} ${} [{}] vs {} - Overall: {} HR Rating: {} Value Rating: {}\n".format(p.name, p.salary, p.teamName, p.oppPitcher.name, round(p.overall, 2), round(p.hrRating, 2), round((p.overall / (float(p.salary) / 1000)), 2))
    output += "\nFirst Basemen to use:\n"
    for p in players['1B']: 
        output += "\t{} ${} [{}] vs {} - Overall: {} HR Rating: {} Value Rating: {}\n".format(p.name, p.salary, p.teamName, p.oppPitcher.name, round(p.overall, 2), round(p.hrRating, 2), round((p.overall / (float(p.salary) / 1000)), 2))
    output += "\nSecond Basemen to use:\n"
    for p in players['2B']: 
        output += "\t{} ${} [{}] vs {} - Overall: {} HR Rating: {} Value Rating: {}\n".format(p.name, p.salary, p.teamName, p.oppPitcher.name, round(p.overall, 2), round(p.hrRating, 2), round((p.overall / (float(p.salary) / 1000)), 2))
    output += "\nThird Basemen to use:\n"
    for p in players['3B']: 
        output += "\t{} ${} [{}] vs {} - Overall: {} HR Rating: {} Value Rating: {}\n".format(p.name, p.salary, p.teamName, p.oppPitcher.name, round(p.overall, 2), round(p.hrRating, 2), round((p.overall / (float(p.salary) / 1000)), 2))
    output += "\nShortstops to use:\n"
    for p in players['SS']: 
        output += "\t{} ${} [{}] vs {} - Overall: {} HR Rating: {} Value Rating: {}\n".format(p.name, p.salary, p.teamName, p.oppPitcher.name, round(p.overall, 2), round(p.hrRating, 2), round((p.overall / (float(p.salary) / 1000)), 2))
    output += "\nOutFielders to use:\n"
    for p in players['OF']: 
        output += "\t{} ${} [{}] vs {} - Overall: {} HR Rating: {} Value Rating: {}\n".format(p.name, p.salary, p.teamName, p.oppPitcher.name, round(p.overall, 2), round(p.hrRating, 2), round((p.overall / (float(p.salary) / 1000)), 2))

    output += "\nHR Ratings, Ranked\n"
    for h in hrList:
        output += "\t{} - HR Rating: {} Overall: {}\n".format(h.name, h.hrRating, h.overall)
    output += "\nStacks to Consider:\n"
    for team in getStacks(hrList):
        output += "{} - Avg Overall: {}\n".format(team['team'], team['averageOverall'])
        for p in team['players']:
            output += "\t[{}] {} ${} - Overall: {} Value Rating: {}\n".format(p.position, p.name, p.salary, round(p.overall, 2), round((p.overall / (float(p.salary) / 1000)), 2))

    summaryFile.write(output)
    summaryFile.close()

def getCellColor(value, standard, over):
    num = value / standard

    maxColor = Colors['Blue'] if over else Colors['Red']
    highColor = Colors['Green'] if over else Colors['Yellow']
    midColor = Colors['White']
    lowColor = Colors['Yellow'] if over else Colors['Green']
    minColor = Colors['Red'] if over else Colors['Blue']

    if num >= 1.15:
        return maxColor
    elif num >= 1.075:
        return highColor
    elif num > .85 and num < 1.075:
        return midColor
    elif num <= .6:
        return minColor
    elif num <= .85:
        return lowColor

def writeSummaryToCSV(hitters, pitchers):
    avgs = leagueAvg.averages
    wb = Workbook()
    #Create Hitters Sheet
    #hitterSheet = wb.create_sheet("Hitters")
    #Create Pitchers Sheet
    pitcherSheet = wb.create_sheet("Pitchers")
    hrHitters = []

    #Append Hitters
    positions = ["C", "1B", "2B", "3B", "SS", "OF"]
    for pos in positions:
        sheet = wb.create_sheet(pos)
        appendRow = ["OU", "Weather", "Pos", "Name", "Team", "Salary", "Hand", "Opp Pitcher", "Overall", "Value", "AB", "Recent wOBA Diff", "Career wOBA Diff", "Recent ISO Diff", "Career ISO Diff", "BABIP", "Career BABIP", "Opp BABIP", "Opp Career BABIP", "Recent FB% Diff", "Career FB% Diff", "Recent HR/FB Diff", "Career HR/FB Diff", "SB%", "HR Rating", "Park HR Factor", "Wind Direction", "Wind Speed", "Humidity", "Temperature", "Order", "FD Pos", "Opp IP", "Opp Career IP"]
        sheet.append(appendRow)
        for hitter in hitters[pos]:
            hrHitters.append(hitter)
            teamOU = hitter.gameCard.getTeamOU(hitter.teamName)

            appendRow = []
            appendRow.append(teamOU)
            appendRow.append(hitter.gameCard.weatherIcon)
            appendRow.append(hitter.position)
            appendRow.append(hitter.name)
            appendRow.append(hitter.teamName)
            appendRow.append(hitter.salary)
            appendRow.append(hitter.handedness)
            appendRow.append(hitter.oppPitcher.name)
            appendRow.append(round(hitter.overall, 2))
            appendRow.append(round((hitter.overall / (float(hitter.salary) / 1000)), 2))

            if hitter.matchupStats == None:
                appendRow.append(0) #AB
                appendRow.append(0) #Recent wOBA Diff
                appendRow.append(0) #Career wOBA Diff
                appendRow.append(0) #Recent ISO Diff
                appendRow.append(0) #Career ISO Diff
                appendRow.append(0) #BABIP
                appendRow.append(0) #Career BABIP
                appendRow.append(0) #Opp BABIP
                appendRow.append(0) #Opp Career BABIP
                appendRow.append(0) #Recent FB% Diff
                appendRow.append(0) #Career FB% Diff
                appendRow.append(0) #Recent HR/FB Diff
                appendRow.append(0) #Career HR/FB Diff
            else:
                appendRow.append(round(hitter.matchupStats['AB'], 3))
                appendRow.append(round( ((hitter.matchupStats['wOBA'] + hitter.oppMatchupStats['wOBA']) / 2) - avgs['wOBA'], 3 ))
                appendRow.append(round( ((hitter.careerMatchupStats['wOBA'] + hitter.careerOppMatchupStats['wOBA']) / 2) - avgs['wOBA'], 3 ))
                appendRow.append(round( ((hitter.matchupStats['ISO'] + hitter.oppMatchupStats['ISO']) / 2) - avgs['ISO'], 3 ))
                appendRow.append(round( ((hitter.careerMatchupStats['ISO'] + hitter.careerOppMatchupStats['ISO']) / 2) - avgs['ISO'], 3 ))
                appendRow.append(round(hitter.matchupStats['BABIP'], 3))
                appendRow.append(round(hitter.careerMatchupStats['BABIP'], 3))
                appendRow.append(round(hitter.oppMatchupStats['BABIP'], 3))
                appendRow.append(round(hitter.careerOppMatchupStats['BABIP'], 3))
                appendRow.append(round( ((hitter.matchupStats['FB%'] + hitter.oppMatchupStats['FB%']) / 2) - avgs['FB'], 3 ))
                appendRow.append(round( ((hitter.careerMatchupStats['FB%'] + hitter.careerOppMatchupStats['FB%']) / 2) - avgs['FB'], 3 ))
                appendRow.append(round( ((hitter.matchupStats['HR/FB'] + hitter.oppMatchupStats['HR/FB']) / 2) - avgs['HR/FB'], 3 ))
                appendRow.append(round( ((hitter.careerMatchupStats['HR/FB'] + hitter.careerOppMatchupStats['HR/FB']) / 2) - avgs['HR/FB'], 3 ))
            try:
                appendRow.append(round(hitter.matchupStats['SB'] / (hitter.matchupStats['1B'] + hitter.matchupStats['HBP'] + hitter.matchupStats['BB']), 2))
            except:
                appendRow.append(0)
            
            appendRow.append(round((hitter.hrRating), 2))
            appendRow.append(hitter.parkFactors['hr'])
            appendRow.append(hitter.gameCard.windDirection)
            appendRow.append(hitter.gameCard.windSpeed)
            appendRow.append(hitter.gameCard.humidity)
            appendRow.append(hitter.gameCard.temperature)

            battingOrder = hitter.gameCard.getPlayerBattingOrder(hitter.name)
            appendRow.append(battingOrder)

            appendRow.append(hitter.FDPos)

            if hitter.matchupStats == None:
                appendRow.append(0) #Opp IP
                appendRow.append(0) #Opp Career IP
            else:
                appendRow.append(hitter.oppMatchupStats['IP'])
                appendRow.append(hitter.careerOppMatchupStats['IP'])

            sheet.append(appendRow)
        sheet.freeze_panes = "A2"

        #Apply Color to Weather - B
        for i in range(sheet.min_row + 1, sheet.max_row):
            cell = sheet["B{}".format(i)]
            c = Colors['Green']
            if cell.value == "Stormy":
                c = Colors['Red']
            elif cell.value == "Rainy":
                c = Colors['Orange']
            elif cell.value == "Cloudy":
                c = Colors['Yellow']
            cell.fill = PatternFill(start_color=c, end_color=c, fill_type = "solid")

        #Apply Colors to wOBA Diff - L, M
        #for i in range(sheet.min_row + 1, sheet.max_row):
        #    cell = sheet["L{}".format(i)]
        #    c = getCellColor(cell.value, avgs['wOBA'], True)
        #    cell.fill = PatternFill(start_color=c, end_color=c, fill_type = "solid")
        #for i in range(sheet.min_row + 1, sheet.max_row):
        #    cell = sheet["M{}".format(i)]
        #    c = getCellColor(cell.value, avgs['wOBA'], True)
        #    cell.fill = PatternFill(start_color=c, end_color=c, fill_type = "solid")
        #Apply Colors to BABIP Diff - P, Q, R, S
        for i in range(sheet.min_row + 1, sheet.max_row):
            cell = sheet["P{}".format(i)]
            c = getCellColor(cell.value, avgs['BABIP'], False)
            cell.fill = PatternFill(start_color=c, end_color=c, fill_type = "solid")
        for i in range(sheet.min_row + 1, sheet.max_row):
            cell = sheet["Q{}".format(i)]
            c = getCellColor(cell.value, avgs['BABIP'], False)
            cell.fill = PatternFill(start_color=c, end_color=c, fill_type = "solid")
        for i in range(sheet.min_row + 1, sheet.max_row):
            cell = sheet["R{}".format(i)]
            c = getCellColor(cell.value, avgs['BABIP'], False)
            cell.fill = PatternFill(start_color=c, end_color=c, fill_type = "solid")
        for i in range(sheet.min_row + 1, sheet.max_row):
            cell = sheet["S{}".format(i)]
            c = getCellColor(cell.value, avgs['BABIP'], False)
            cell.fill = PatternFill(start_color=c, end_color=c, fill_type = "solid")

    appendRow = ["OU", "Weather", "Name", "Team", "Salary", "Hand", "Opp Team", "Overall", "Value", "IP", "K% vs L", "K% vs R", "Opp K%", "wOBA", "Career wOBA", "Opp wOBA", "ISO", "Opp ISO", "BABIP", "Opp BABIP", "xFIP", "Park HR Factor", "Wind Direction", "Wind Speed", "Humidity", "Temperature"]
    pitcherSheet.append(appendRow)
    for pitcher in pitchers:
        teamOU = pitcher.gameCard.getTeamOU(pitcher.oppTeamName)

        appendRow = []
        appendRow.append(teamOU)
        appendRow.append(pitcher.gameCard.weatherIcon)
        appendRow.append(pitcher.name)
        appendRow.append(pitcher.teamName)
        appendRow.append(pitcher.salary)
        appendRow.append(pitcher.handedness)
        appendRow.append(pitcher.oppTeamName)
        appendRow.append(round(pitcher.overall, 2))
        appendRow.append(round((hitter.overall / (float(pitcher.salary) / 1000)), 2))
        appendRow.append(pitcher.stats['vsL']['IP'] + pitcher.stats['vsR']['IP'])
        appendRow.append(round(pitcher.kRate['vsL'], 2))
        appendRow.append(round(pitcher.kRate['vsR'], 2))
        appendRow.append(round(pitcher.kRate['opp'], 2))
        try:
            appendRow.append(round((pitcher.stats['vsL']['wOBA'] + pitcher.stats['vsR']['wOBA']) / 2, 3))
        except:
            print("{} Missing wOBA".format(pitcher.name))
            appendRow.append(0)
        try:
            appendRow.append(round((pitcher.stats['careerVsL']['wOBA'] + pitcher.stats['careerVsR']['wOBA']) / 2, 3))
        except:
            print("{} Missing Career wOBA".format(pitcher.name))
            appendRow.append(0)
        appendRow.append(round(pitcher.oppMatchupStats['wOBA'], 3))
        try:
            appendRow.append(round((pitcher.stats['vsL']['ISO'] + pitcher.stats['vsR']['ISO']) / 2, 3))
        except:
            print("{} Missing ISO".format(pitcher.name))
            appendRow.append(0)
        appendRow.append(round(pitcher.oppMatchupStats['ISO'], 3))
        try:
            appendRow.append(round((pitcher.stats['vsL']['BABIP'] + pitcher.stats['vsR']['BABIP']) / 2, 3))
        except:
            print("{} Missing BABIP".format(pitcher.name))
            appendRow.append(0)
        appendRow.append(round(pitcher.oppMatchupStats['BABIP'], 3))
        try:
            appendRow.append(round((pitcher.stats['vsL']['xFIP'] + pitcher.stats['vsR']['xFIP']) / 2, 2))
        except:
            print("{} Missing xFIP".format(pitcher.name))
            appendRow.append(0)
        appendRow.append(pitcher.parkFactors['hr'])
        appendRow.append(pitcher.gameCard.windDirection)
        appendRow.append(pitcher.gameCard.windSpeed)
        appendRow.append(pitcher.gameCard.humidity)
        appendRow.append(pitcher.gameCard.temperature)
        
        pitcherSheet.append(appendRow)

    pitcherSheet.freeze_panes = "A2"

    #Apply Color to Weather - B
    for i in range(pitcherSheet.min_row + 1, pitcherSheet.max_row):
        cell = pitcherSheet["B{}".format(i)]
        c = Colors['Green']
        if cell == "Stormy":
            c = Colors['Red']
        elif cell == "Rainy":
            c = Colors['Orange']
        elif cell == "Cloudy":
            c = Colors['Yellow']
        cell.fill = PatternFill(start_color=c, end_color=c, fill_type = "solid")

    #Apply Colors to wOBA Diff - N, O
    #for i in range(pitcherSheet.min_row + 1, pitcherSheet.max_row):
    #    cell = pitcherSheet["N{}".format(i)]
    #    c = getCellColor(cell.value, avgs['wOBA'], False)
    #    cell.fill = PatternFill(start_color=c, end_color=c, fill_type = "solid")
    #for i in range(pitcherSheet.min_row + 1, pitcherSheet.max_row):
    #    cell = pitcherSheet["O{}".format(i)]
    #    c = getCellColor(cell.value, avgs['wOBA'], True)
    #    cell.fill = PatternFill(start_color=c, end_color=c, fill_type = "solid")
    #Apply Colors to BABIP Diff - R, S
    #for i in range(pitcherSheet.min_row + 1, pitcherSheet.max_row):
    #    cell = pitcherSheet["R{}".format(i)]
    #    c = getCellColor(cell.value, avgs['BABIP'], False)
    #    cell.fill = PatternFill(start_color=c, end_color=c, fill_type = "solid")
    #for i in range(pitcherSheet.min_row + 1, pitcherSheet.max_row):
    #    cell = pitcherSheet["S{}".format(i)]
    #    c = getCellColor(cell.value, avgs['BABIP'], False)
    #    cell.fill = PatternFill(start_color=c, end_color=c, fill_type = "solid")

    #Add sheet for HR list
    sheet = wb.create_sheet("HR")
    hrHitters.sort(key=lambda x: x.hrRating, reverse=True)

    header = ['Weather', 'Name', 'Pos', 'HR Rating']
    sheet.append(header)
    dataRow = []

    count = 0
    for hrHitter in hrHitters:
        if count >= 15:
            break
        dataRow = []
        dataRow.append(hrHitter.gameCard.weatherIcon)
        dataRow.append(hrHitter.name)
        dataRow.append(hrHitter.position)
        dataRow.append(hrHitter.hrRating)

        sheet.append(dataRow)

        count += 1
    sheet.freeze_panes = "A2"

    wb.save("Summary.xlsx")