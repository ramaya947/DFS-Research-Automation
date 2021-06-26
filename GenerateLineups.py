from openpyxl import Workbook, load_workbook

def generateLineups():
    wb = load_workbook("Summary.xlsx")
    keys = ['Pitchers', 'C', '1B', '2B', '3B', 'SS', 'OF']
    excelHeaderKey = {
        'Fanduel ID Pitcher': 'AA',
        'Fanduel ID Hitter': 'AC',
        'Name Pitcher': 'C',
        'Name Hitter': 'D' ,
        'Salary Pitcher': 'E',
        'Salary Hitter': 'F',
        'Fanduel Position Hitter': 'Z',
        'Team Name Pitcher': 'D',
        'Team Name Hitter': 'E'
    }
    players = {
        'Pitchers': [],
        'C/1B': [],
        '2B': [],
        '3B': [],
        'SS': [],
        'OF': [],
        'Utils': []
    }
    greenRGB = "FF00FF00"

    #Get All Selected Players
    for sheetName in keys:
        sheet = wb.get_sheet_by_name(sheetName)
        working = True
        count = 2 #This is the first non-header index
        nameColKey = excelHeaderKey['Name Pitcher'] if (sheetName == 'Pitchers') else excelHeaderKey['Name Hitter']
        fanduelIDColKey = excelHeaderKey['Fanduel ID Pitcher'] if (sheetName == 'Pitchers') else excelHeaderKey['Fanduel ID Hitter']
        salaryColKey = excelHeaderKey['Salary Pitcher'] if (sheetName == 'Pitchers') else excelHeaderKey['Salary Hitter']
        teamNameColKey = excelHeaderKey['Team Name Pitcher'] if (sheetName == 'Pitchers') else excelHeaderKey['Team Name Hitter']

        while working:
            nameCell = sheet['{}{}'.format(nameColKey, count)]
            if nameCell.value == "" or nameCell.value == None:
                working = None
                continue

            fill = nameCell.fill.bgColor.rgb
            if fill != greenRGB:
                count += 1
                continue

            fanduelIDCell = sheet['{}{}'.format(fanduelIDColKey, count)]

            salaryCell = sheet['{}{}'.format(salaryColKey, count)]

            teamNameCell = sheet['{}{}'.format(teamNameColKey, count)]
            
            fdPosition = 'P' if sheetName == 'Pitchers' else sheet['{}{}'.format(excelHeaderKey['Fanduel Position Hitter'], count)]

            positions = []

            if sheetName == 'Pitchers':
                positions = ['Pitchers']
            else:
                positions = fdPosition.value.split('/')

            for pos in positions:
                if pos == 'C' or pos == '1B':
                    pos = 'C/1B'

                players[pos].append({
                    'Fanduel ID': fanduelIDCell.value,
                    'Name': nameCell.value,
                    'Salary': float(salaryCell.value),
                    'Team Name': teamNameCell.value
                })

            if sheetName != 'Pitchers':
                players['Utils'].append({
                    'Fanduel ID': fanduelIDCell.value,
                    'Name': nameCell.value,
                    'Salary': float(salaryCell.value),
                    'Team Name': teamNameCell.value
                })

            count += 1

    keys = ['Pitchers', 'C/1B', '2B', '3B', 'SS', 'OF']

    #Generate Unique Lineups
    stop = False
    comboCount = 0
    lineupCount = 0
    lineups = {}
    stackKeys = []
    for pitcher in players['Pitchers']:
        if stop:
            break
        for first in players['C/1B']:
            if stop:
                break
            for second in players['2B']:
                if stop:
                    break
                for third in players['3B']:
                    if stop:
                        break
                    for ss in players['SS']:
                        if stop:
                            break
                        for OF1 in players['OF']:
                            if stop:
                                break
                            if comboCount > 5:
                                comboCount = 0 
                                break
                            OF2List = removeFromList(OF1, players['OF'][:])
                            for OF2 in OF2List:
                                if stop:
                                    break
                                if comboCount > 5: 
                                    break
                                OF3List = removeFromList(OF2, OF2List[:])
                                for OF3 in OF3List:
                                    if stop:
                                        break
                                    if comboCount > 5: 
                                        break
                                    #Remove all other players from a copy of the Utils List
                                    utilsFiltered = players['Utils'][:]
                                    removePlayersFromList([first, second, third, ss, OF1, OF2, OF3], utilsFiltered)
                                    
                                    for util in utilsFiltered:
                                        if comboCount > 5: 
                                            break

                                        lineup = {
                                            'Pitcher': pitcher,
                                            'C/1B': first,
                                            '2B': second,
                                            '3B': third,
                                            'SS': ss,
                                            'OF1': OF1,
                                            'OF2': OF2,
                                            'OF3': OF3,
                                            'Util': util
                                        }

                                        if (getTotalSalary([pitcher, first, second, third, ss, OF1, OF2, OF3, util]) > 35000) or alreadyMade(lineup, lineups) or oversizeStack(lineup):
                                            continue

                                        stackInfo = getLineupStack(lineup)

                                        if stackInfo['Team'] not in lineups:
                                            lineups[stackInfo['Team']] = [lineup]
                                            stackKeys.append(stackInfo['Team'])
                                        else:
                                            lineups[stackInfo['Team']].append(lineup)

                                        comboCount += 1
                                        lineupCount += 1
                                        print ("Total lineups created: {}".format(lineupCount), end="\r")

                                        if (lineupCount >= 1000):
                                            stop = True
                                            break

    lineupWB = Workbook()
    sheet = lineupWB.create_sheet("Lineups")

    headerRow = ['P', 'C/1B', '2B', '3B', 'SS', 'OF', 'OF', 'OF', 'UTIL']
    sheet.append(headerRow)

    count = 0
    stackIndex = 0
    while count < 150:
        key = stackKeys[stackIndex]
        stackIndex += 1
        if (stackIndex >= len(stackKeys)):
            stackIndex = 0

        if (len(lineups[key]) == 0):
            nothingLeft = True
            for k in stackKeys:
                if (len(lineups[key]) != 0):
                    nothingLeft = False
            if nothingLeft:
                break
            continue
        l = lineups[key].pop(0)

        appendRow = []
        appendRow.append(l['Pitcher']['Fanduel ID'])
        appendRow.append(l['C/1B']['Fanduel ID'])
        appendRow.append(l['2B']['Fanduel ID'])
        appendRow.append(l['3B']['Fanduel ID'])
        appendRow.append(l['SS']['Fanduel ID'])
        appendRow.append(l['OF1']['Fanduel ID'])
        appendRow.append(l['OF2']['Fanduel ID'])
        appendRow.append(l['OF3']['Fanduel ID'])
        appendRow.append(l['Util']['Fanduel ID'])

        printLineup(l, key)

        sheet.append(appendRow)

        count += 1

    lineupWB.save("Lineups.xlsx")

def printLineup(l, stackName):
    output = "Stack: ({}) - [P] {} [C/1B] {} [2B] {} [3B] {} [SS] {} [OF] {}, {}, {} [Util] {}".format(stackName, l['Pitcher']['Name'], l['C/1B']['Name'], l['2B']['Name'], l['3B']['Name'], l['SS']['Name'], l['OF1']['Name'], l['OF2']['Name'], l['OF3']['Name'], l['Util']['Name'])

    print(output)

def oversizeStack(lineup):
    teams = {}
    keys = ['C/1B', '2B', '3B', 'SS', 'OF1', 'OF2', 'OF3', 'Util']

    for key in keys:
        p = lineup[key]

        if p['Team Name'] not in teams:
            teams[p['Team Name']] = 1
        else:
            teams[p['Team Name']] = teams[p['Team Name']] + 1

        if teams[p['Team Name']] > 4:
            return True
    
    return False

def getLineupStack(lineup):
    teams = {}
    teamKeys = []
    keys = ['C/1B', '2B', '3B', 'SS', 'OF1', 'OF2', 'OF3', 'Util']

    for key in keys:
        p = lineup[key]

        if p['Team Name'] not in teams:
            teams[p['Team Name']] = 1
            teamKeys.append(p['Team Name'])
        else:
            teams[p['Team Name']] = teams[p['Team Name']] + 1

    primaryStack = ''
    primaryStackValue = None
    for key in teamKeys:
        if primaryStack == '' or primaryStackValue == None:
            primaryStack = key
            primaryStackValue = teams[key]
        else:
            if primaryStackValue < teams[key]:
                primaryStack = key
                primaryStackValue = teams[key]

    return { 'Team': primaryStack, 'Stack Count': primaryStackValue }

def removeFromList(toRemove, theList):
    for x in range(0, len(theList)):
        p = theList[x]

        if toRemove['Fanduel ID'] == p['Fanduel ID']:
            del theList[x]
            break

    return theList

def removePlayersFromList(playersToRemove, theList):
    for p1 in playersToRemove:
        for x in range(0, len(theList)):
            p2 = theList[x]

            if p1['Fanduel ID'] == p2['Fanduel ID']:
                del theList[x]
                break

    return theList

def getTotalSalary(players):
    total = 0

    for p in players:
        total += p['Salary']

    return total

def alreadyMade(curLineup, lineups):
    for x in lineups:
        y = lineups[x]
        for l in y:
            ids = [
                l['Pitcher']['Fanduel ID'],
                l['C/1B']['Fanduel ID'],
                l['2B']['Fanduel ID'],
                l['3B']['Fanduel ID'],
                l['SS']['Fanduel ID'],
                l['OF1']['Fanduel ID'],
                l['OF2']['Fanduel ID'],
                l['OF3']['Fanduel ID'],
                l['Util']['Fanduel ID']
            ]

            #Remove all matching
            try:
                ids.remove(curLineup['Pitcher']['Fanduel ID'])
            except:
                pass
            try:
                ids.remove(curLineup['C/1B']['Fanduel ID'])
            except:
                pass
            try:
                ids.remove(curLineup['2B']['Fanduel ID'])
            except:
                pass
            try:
                ids.remove(curLineup['3B']['Fanduel ID'])
            except:
                pass
            try:
                ids.remove(curLineup['SS']['Fanduel ID'])
            except:
                pass
            try:
                ids.remove(curLineup['OF1']['Fanduel ID'])
            except:
                pass
            try:
                ids.remove(curLineup['OF2']['Fanduel ID'])
            except:
                pass
            try:
                ids.remove(curLineup['OF3']['Fanduel ID'])
            except:
                pass
            try:
                ids.remove(curLineup['Util']['Fanduel ID'])
            except:
                    pass

            if len(ids) == 0:
                #This is not unique, return True to not add
                return True

    #If you got here, that means it is unique
    return False

generateLineups()