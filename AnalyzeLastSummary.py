from bs4 import BeautifulSoup
import requests
from datetime import datetime, timedelta
import re
import pandas as pd
from openpyxl import Workbook

def analyze():
    wb = Workbook()
    hitterSheetKeys = ['Pitchers', 'C', '1B', '2B', '3B', 'SS', 'OF']
    excelData = None
    optimalPerformance = {
        'Pitchers': [], 
        'C': [],
        '1B': [],
        '2B': [],
        '3B': [],
        'SS': [],
        'OF': [] 
    }

    namesToScores = getNamesToScores()
    hittersForStackResults = {}

    for key in hitterSheetKeys:
        sheet = wb.create_sheet(key)
        excelData = pd.read_excel('SummaryLast.xlsx', sheet_name=key)

        if key == 'Pitchers':
            namesToCheck = excelData['Name'].tolist()
            overalls = excelData['Overall'].tolist()
            salaries = excelData['Salary'].tolist()
            overUnders = excelData['OU'].tolist()
            kvsLs = excelData['K% vs L'].tolist()
            kvsRs = excelData['K% vs R'].tolist()
            oppK = excelData['Opp K%'].tolist()
            recwOBA = excelData['wOBA'].tolist()
            carwOBA = excelData['Career wOBA'].tolist()
            oppwOBA = excelData['Opp wOBA'].tolist()
            iso = excelData['ISO'].tolist()
            oppISO = excelData['Opp ISO'].tolist()
            babip = excelData['BABIP'].tolist()
            oppBabip = excelData['Opp BABIP'].tolist()
            xFIP = excelData['xFIP'].tolist()

            header = ["OU", "Name", "Salary", "Overall", "K% vs L", "K% vs R", "Opp K%", "wOBA", "Career wOBA", "Opp wOBA", "ISO", "Opp ISO", "BABIP", "Opp BABIP", "xFIP", "Points"]
            sheet.append(header)

            count = 0
            for name in namesToCheck:
                dataRow = []

                score = None
                try:
                    score = namesToScores[name]
                except:
                    count += 1
                    continue

                dataRow.append(overUnders[count])
                dataRow.append(name)
                dataRow.append(salaries[count])
                dataRow.append(overalls[count])
                dataRow.append(kvsLs[count])
                dataRow.append(kvsRs[count])
                dataRow.append(oppK[count])
                dataRow.append(recwOBA[count])
                dataRow.append(carwOBA[count])
                dataRow.append(oppwOBA[count])
                dataRow.append(iso[count])
                dataRow.append(oppISO[count])
                dataRow.append(babip[count])
                dataRow.append(oppBabip[count])
                dataRow.append(xFIP[count])
                dataRow.append(score)
                
                sheet.append(dataRow)

                try:
                    if len(optimalPerformance[key]) != 0:
                        if float(score) > float(optimalPerformance[key][-1]['Score']):
                            optimalPerformance[key].append({'Name': name, 'Score': score, 'Salary': salaries[count]})
                    else:
                        optimalPerformance[key].append({'Name': name, 'Score': score, 'Salary': salaries[count]})
                except:
                    pass

                count += 1

        else:
            namesToCheck = excelData['Name'].tolist()
            overalls = excelData['Overall'].tolist()
            salaries = excelData['Salary'].tolist()
            overUnders = excelData['OU'].tolist()
            recwOBADiff = excelData['Recent wOBA Diff'].tolist()
            carwOBADiff = excelData['Career wOBA Diff'].tolist()
            recISODiff = excelData['Recent ISO Diff'].tolist()
            carISODiff = excelData['Career ISO Diff'].tolist()
            BABIP = excelData['BABIP'].tolist()
            carBABIP = excelData['Career BABIP'].tolist()
            oppBABIP = excelData['Opp BABIP'].tolist()
            oppCarBABIP = excelData['Opp Career BABIP'].tolist()
            recFBDiff = excelData['Recent FB% Diff'].tolist()
            carFBDiff = excelData['Career FB% Diff'].tolist()
            recHRFBDiff = excelData['Recent HR/FB Diff'].tolist()
            carHRFBDiff = excelData['Career HR/FB Diff'].tolist()
            hrRating = excelData['HR Rating'].tolist()

            header = ["OU", "Name", "Salary", "Overall", "Recent wOBA Diff", "Career wOBA Diff", "Recent ISO Diff", "Career ISO Diff", "BABIP", "Career BABIP", "Opp BABIP", "Opp Career BABIP", "Recent FB% Diff", "Career FB% Diff", "Recent HR/FB Diff", "Career HR/FB Diff", "HR Rating", "Points"]
            sheet.append(header)

            count = 0
            for name in namesToCheck:
                dataRow = []

                score = None
                try:
                    score = namesToScores[name]
                except:
                    count += 1
                    continue

                dataRow.append(overUnders[count])
                dataRow.append(name)
                dataRow.append(salaries[count])
                dataRow.append(overalls[count])
                dataRow.append(recwOBADiff[count])
                dataRow.append(carwOBADiff[count])
                dataRow.append(recISODiff[count])
                dataRow.append(carISODiff[count])
                dataRow.append(BABIP[count])
                dataRow.append(carBABIP[count])
                dataRow.append(oppBABIP[count])
                dataRow.append(oppCarBABIP[count])
                dataRow.append(recFBDiff[count])
                dataRow.append(carFBDiff[count])
                dataRow.append(recHRFBDiff[count])
                dataRow.append(carHRFBDiff[count])
                dataRow.append(hrRating[count])
                dataRow.append(score)
                
                sheet.append(dataRow)

                try:
                    if len(optimalPerformance[key]) != 0:
                        if float(score) > float(optimalPerformance[key][-1]['Score']):
                            optimalPerformance[key].append({'Name': name, 'Score': score, 'Salary': salaries[count]})
                    else:
                        optimalPerformance[key].append({'Name': name, 'Score': score, 'Salary': salaries[count]})
                except:
                    pass

                count += 1
        
        sheet.freeze_panes = "A2"

    sheet = wb.create_sheet("Optimal Lineup")

    header = ["Position", "Name", "Salary", "Score"]
    sheet.append(header)

    #Find Best Pitcher
    dataRow = []
    player = optimalPerformance['Pitchers'].pop()

    dataRow.append('Pitcher')
    dataRow.append(player['Name'])
    dataRow.append(player['Salary'])
    try:
        dataRow.append(float(player['Score']))
    except:
        dataRow.append(0)
    sheet.append(dataRow)

    #Find Best C/1B
    dataRow = []
    try:
        if float(optimalPerformance['C'][-1]['Score']) > float(optimalPerformance['1B'][-1]['Score']):
            player = optimalPerformance['C'].pop()
        else:
            player = optimalPerformance['1B'].pop()
    except:
        player = optimalPerformance['C'].pop()

    dataRow.append('C/1B')
    dataRow.append(player['Name'])
    dataRow.append(player['Salary'])
    try:
        dataRow.append(float(player['Score']))
    except:
        dataRow.append(0)
    sheet.append(dataRow)

    #Find Best 2B
    dataRow = []
    player = optimalPerformance['2B'].pop()

    dataRow.append('2B')
    dataRow.append(player['Name'])
    dataRow.append(player['Salary'])
    try:
        dataRow.append(float(player['Score']))
    except:
        dataRow.append(0)
    sheet.append(dataRow)

    #Find Best 3B
    dataRow = []
    player = optimalPerformance['3B'].pop()

    dataRow.append('3B')
    dataRow.append(player['Name'])
    dataRow.append(player['Salary'])
    try:
        dataRow.append(float(player['Score']))
    except:
        dataRow.append(0)
    sheet.append(dataRow)

    #Find Best SS
    dataRow = []
    player = optimalPerformance['SS'].pop()

    dataRow.append('SS')
    dataRow.append(player['Name'])
    dataRow.append(player['Salary'])
    try:
        dataRow.append(float(player['Score']))
    except:
        dataRow.append(0)
    sheet.append(dataRow)

    #Find Best OF x 3
    for x in range(0, 3):
        dataRow = []
        try:
            player = optimalPerformance['OF'].pop()
        except:
            continue

        dataRow.append('OF')
        dataRow.append(player['Name'])
        dataRow.append(player['Salary'])
        try:
            dataRow.append(float(player['Score']))
        except:
            dataRow.append(0)
        sheet.append(dataRow)

    #Find Best UTIL
    dataRow = []
    player = None
    keys = ['C', '1B', '2B', '3B', 'SS', 'OF']
    for key in keys:
        if player == None:
            try:
                player = optimalPerformance[key].pop()
            except:
                continue
        else:
            try:
                if float(player['Score']) < float(optimalPerformance[key][-1]['Score']):
                    player = optimalPerformance[key].pop()
            except:
                pass

    dataRow.append('Util')
    dataRow.append(player['Name'])
    dataRow.append(player['Salary'])
    try:
        dataRow.append(float(player['Score']))
    except:
        dataRow.append(0)
    sheet.append(dataRow)

    sheet.freeze_panes = "A2"

    #Give Results for Stacks
    excelData = pd.read_excel('SummaryLast.xlsx', sheet_name="Stacks", header=None, index_col=0)
    colZero = excelData[0].tolist()
    colTwoPosition = excelData[2].tolist()
    colThreeNames = excelData[3].tolist()

    count = 0
    
    
    wb.save("SummaryAnalyzed {}.xlsx".format(getDateForFileName()))

def getNamesToScores():
    URL = "https://rotogrinders.com/lineups/mlb?date={}&site=fanduel"
    url = URL.format(getDate())
    scores = {}

    data = requests.get(url)
    soup = BeautifulSoup(data.text)
    cards = soup.find_all("div", {"class": "blk crd lineup"})

    for card in cards:
        playerListSoups = card.find_all("li", {"class": "player"})
        pitcherListSoups = card.find_all("div", {"class": "pitcher players"})
        
        for playerSoup in playerListSoups:
            infoSoup = playerSoup.find("div", {"class": "info"})
            name = infoSoup.find("a", {"class": "player-popup"}).text
            score = infoSoup.find("span", {"class": "fpts actual"}).text

            scores[name] = score

        for pitcherSoup in pitcherListSoups:
            name = pitcherSoup.find("a", {"class": "player-popup"}).text
            score = pitcherSoup.find("span", {"class": "fpts actual"}).text

            scores[name] = score

    return scores

def getDate():
    dt = datetime.now() + timedelta(days=-1)

    day = ""
    if dt.day < 10:
        day = "0{}".format(dt.day)
    else:
        day = dt.day

    month = ""
    if dt.month < 10:
        month = "0{}".format(dt.month)
    else:
        month = dt.month
    
    year = dt.year
    
    date = "{}-{}-{}".format(year, month, day)

    return date

def getDateForFileName():
    dt = datetime.now() + timedelta(days=-1)

    day = dt.day
    month = dt.month
    year = dt.year
    
    date = "{}-{}-{}".format(month, day, year)

    return date

analyze()