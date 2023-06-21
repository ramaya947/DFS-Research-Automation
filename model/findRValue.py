import numpy as np
from sklearn.linear_model import LinearRegression
from openpyxl import load_workbook
import pocketbase.PocketbaseProxy as pocketbase

def findRValue(x, y):
    x = np.array(x).reshape((-1, 1))
    y = np.array(y)

    model = LinearRegression().fit(x, y)
    return model.score(x, y)

# The source xlsx file is named as source.xlsx
wb=load_workbook("Analysis.xlsx")

rValues = {}
for sheetKey in wb.sheetnames:
    sheet = wb[sheetKey]
    fptsColumn = sheet['AR']
    fpts = []
    isHeaderRow = True
    for x in range(len(fptsColumn)):
        if isHeaderRow:
            isHeaderRow = False
            continue
        fpts.append(fptsColumn[x].value) 

    rValues[sheetKey] = []
    for x in sheet.columns:
        values = []
        isHeaderRow = True
        statType = ''
        for cell in x:
            if isHeaderRow:
                isHeaderRow = False
                statType = cell.value
                continue
            values.append(cell.value)
        isHeaderRow = True

        rValue = findRValue(values, fpts)
        rValues[sheetKey].append(
            {
                'position': sheetKey,
                'stat': statType,
                'rValue': round(rValue, 4) * 100
            }
        )
        #print('R Value for {} is {}'.format(statType, rValue))

for sheetKey in wb.sheetnames:
    rValues[sheetKey].sort(key=lambda x: x['rValue'], reverse=True)
    for val in rValues[sheetKey]:
        print(val)
        pocketbase.addRValue('rValues', val)
