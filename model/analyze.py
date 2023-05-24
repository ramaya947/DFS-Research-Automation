import pocketbase.PocketbaseProxy as pocketbase
from openpyxl import Workbook, utils

# TODO: Call getAllPlayers
players = pocketbase.getAllPlayers('players')

wb = Workbook()

# TODO: Create workbook sheets for all positions
positions = ['P', 'C', '1B', '2B', '3B', 'SS', 'OF']
dict = {}
for pos in positions:
	dict[pos] = {
		'sheet': wb.create_sheet(pos),
		'hasHeaders': False
	}
dict['General'] = {
	'sheet': wb.create_sheet('General'),
	'hasHeaders': False
}

count = 0
for player in players:
	performances = pocketbase.getAllPerformanceRecords('performances', 'fid=\'{}\''.format(player['fid']))

	for performance in performances:
		if performance['positions'] == '':
			continue

		positions = performance['positions'].split('/')
		sheets = []
		for pos in positions:
			sheets.append(dict[pos])
		sheets.append(dict['General'])

		headerRow = []
		statRow = []
		# Add all headers if needed, add stats for performance
		performanceStats = performance['stats']
		for key in performanceStats.keys():
			stat = performanceStats[key]
			if isinstance(stat, str) or key == 'Season':
				continue
			headerRow.append('{}'.format(key))
			statRow.append(stat)

		headerRow.append('FPTS')
		if not isinstance(performance['fpts'], int) and not isinstance(performance['fpts'], float):
			continue
		statRow.append(performance['fpts'])
		for sheet in sheets:
			# After first pass through a performance, headers should be added and can be set to True
			if not sheet['hasHeaders']:
				sheet['sheet'].append(headerRow)
				sheet['hasHeaders'] = True

			sheet['sheet'].append(statRow)

	count += 1
	print('Player {} out of {}'.format(count + 1, len(players)), flush=True)
wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
wb.save("Analysis.xlsx")